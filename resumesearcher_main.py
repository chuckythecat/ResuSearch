"""TODO:
фичи:
диалоговое окно с подтверждением удаления файла
диалоговое окно с выбором файла базы данных если не найден файл бд в папке с прогой


говнокодинг:
привязка в таблице идет по имени и фамилии 
(один из людей у которого есть тезка по имени и фамилии удаляется из таблицы)

записывать всю информацию о людях в один словарь
количество строк которые можно заполнить ограничено
"""

from watchgod import watch
from PyQt5 import QtCore
import ctypes
import csv
import fix_qt_import_error
from PyQt5 import QtWidgets
import resumesearcher_ui
from sys import argv
from os import remove, startfile, getcwd
import openpyxl
from ntpath import basename

dbfilename = "resumaker_database.rdb"

excelbook = openpyxl.Workbook()
excelsheet = excelbook.create_sheet("База данных резюмейкер", 0)

excelsheet['A1'] = "Имя"
excelsheet['B1'] = "Фамилия"
excelsheet['C1'] = "Должность"
excelsheet['D1'] = "Возраст"
excelsheet['E1'] = "Уровень образования"

excelsheet.column_dimensions['A'].width = 10
excelsheet.column_dimensions['B'].width = 10
excelsheet.column_dimensions['C'].width = 13
excelsheet.column_dimensions['D'].width = 3
excelsheet.column_dimensions['E'].width = 20

personsdict, pathsdict = {}, {}
jobslist = []

myappid = u'mycompany.myproduct.subproduct.version'

ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)


def getdictfromfile(filename):
    try:
        reader = csv.reader(open(filename, 'r', encoding="utf-8"))
    except FileNotFoundError:
        print("file not found")
        exit(0)

    for line in reader:
        name = line[0]
        second_name = line[1]
        post = line[2]
        age = int(line[3])
        edu_level = line[4]
        resume_path = line[5]
        personsdict[(name, second_name)] = [name, second_name, post, age, edu_level]
        pathsdict[(name, second_name)] = resume_path
        global jobslist
        jobslist.append(post)
        jobslist = list(set(jobslist))


def export():
    for x, (persons, info) in enumerate(personsdict.items()):
        for y, item in enumerate(info):
            excelsheet.cell(row=(x + 2), column=(y + 1), value=item)
    excelbook.save("resumaker_database.xlsx")
    startfile('resumaker_database.xlsx', 'open')


class Worker(QtCore.QThread):

    database_updated = QtCore.pyqtSignal(object)

    @QtCore.pyqtSlot()
    def run(self):
        print("Watchdog thread started")
        for changes in watch(getcwd()):
            print(changes)
            if changes and basename(next(iter(changes))[1]) == dbfilename:
                self.database_updated.emit(True)
                print("database changed")


class ResumeSearcher(QtWidgets.QMainWindow, resumesearcher_ui.Ui_MainWindow):
    def __init__(self):
        # setup
        super().__init__()
        self.setupUi(self)

        # connect button and combobox signals to class methods
        self.editresumeButton.clicked.connect(self.edit)
        self.deleteButton.clicked.connect(self.delete)
        self.excelexportButton.clicked.connect(export)
        self.comboBox.currentIndexChanged.connect(self.on_index_change)

        # initialise multithreading
        self.threadpool = QtCore.QThreadPool()

        # exit if only 1 thread available, else print amount of available threads
        if self.threadpool.maxThreadCount() < 2:
            exit("lol potato pc")
        else:
            print("Multithreading with maximum of %d threads" % self.threadpool.maxThreadCount())

        # start second thread watching for database file changes
        # worker = Worker()
        # worker.database_updated.connect(self.on_database_update)
        # worker.start()

        # for line in reader:
        #     name = line[0]
        #     second_name = line[1]
        #     post = line[2]
        #     age = int(line[3])
        #     edu_level = line[4]
        #     resume_path = line[5]
        #     personsdict[(name, second_name)] = [name, second_name, post, age, edu_level]
        #     pathsdict[(name, second_name)] = resume_path
        #     jobslist1.append(post)
        #     global jobslist
        #     jobslist = list(set(jobslist1))
        getdictfromfile(dbfilename)

        # initialise combobox
        for i in jobslist:
            self.comboBox.addItem(i)
        self.rebuildtable(personsdict)

    def on_database_update(self):
        personsdict.clear()
        pathsdict.clear()
        jobslist.clear()
        getdictfromfile(dbfilename)
        self.rebuildtable(personsdict)

    def rebuildtable(self, rebuildabledict):
        self.peopletableWidget.setRowCount(len(rebuildabledict))

        for i, (persons, info) in enumerate(rebuildabledict.items()):
            # print(i, persons)
            for x, item in enumerate(info):
                # print("\t", x, items)
                qtableitem = QtWidgets.QTableWidgetItem()
                self.peopletableWidget.setItem(i, x, qtableitem)
                self.peopletableWidget.item(i, x).setText(str(item))

    def getpersonsinfo(self):
        return (self.peopletableWidget.item(self.peopletableWidget.currentRow(), 0).text(),
                self.peopletableWidget.item(self.peopletableWidget.currentRow(), 1).text())

    def edit(self):
        startfile(pathsdict[self.getpersonsinfo()], 'open')
        # print(personsdict)

    def delete(self):
        self.peopletableWidget.setSortingEnabled(False)
        personsinfo = self.getpersonsinfo()
        personsdict.pop(personsinfo)
        remove(pathsdict[personsinfo])
        pathsdict.pop(personsinfo)

        writer = csv.writer(open(dbfilename, "w"), delimiter=",")
        data = []
        for persons, info in personsdict.items():
            data.append(info)
        print(data)
        for line in data:
            writer.writerow(line)

        self.rebuildtable(personsdict)
        self.peopletableWidget.setSortingEnabled(True)

    def on_index_change(self):
        self.peopletableWidget.setSortingEnabled(False)
        if self.comboBox.currentText() == "":
            self.rebuildtable(personsdict)
        else:
            sorteddict = personsdict.copy()
            for a, b in list(sorteddict.items()):
                if b[2] != self.comboBox.currentText():
                    sorteddict.pop(a)
            self.rebuildtable(sorteddict)
        self.peopletableWidget.setSortingEnabled(True)


def main():
    app = QtWidgets.QApplication(argv)
    window = ResumeSearcher()
    window.show()
    app.exec_()


if __name__ == '__main__':
    main()
